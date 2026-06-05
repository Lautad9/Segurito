"""
Generador de AR/ATS: rellena la plantilla .xlsm de YPF a partir de una
especificación de tarea (JSON), preservando fórmulas y macros (.xlsm).

Hojas de ENTRADA que escribimos:
  - INSTALACIÓN: título, instalación, versión
  - FASES DE TAREA: por fase -> título, detalle, personas, herramientas, maquinaria, izaje
  - RIESGOS: por fase -> nombres de riesgo (col D) + barrera principal (col E)
Hojas CALCULADAS (no se tocan, se recalculan solas):
  - IPER (matriz), BARRERAS, EVALUACION (scoring Fine), IMPRIMIR (salida final)
"""
import openpyxl, warnings, json, shutil, sys
try:
    from extraer_catalogo import norm
except ImportError:
    from engine.extraer_catalogo import norm
warnings.filterwarnings("ignore")

# Anclajes fijos de la plantilla (detectados desde las fórmulas del original)
FASES_TITULO_ROW = [27, 31, 39, 45, 53, 60]          # col C = título de fase
RIESGOS_RANGE = [(27,58),(59,91),(93,129),(131,160),(164,191),(195,222)]  # D-range por fase (matriz IPER)
FASES_FIN_BLOQUE = 72   # última fila del bloque de fases (antes de EN CASO DE ACCIDENTE en fila 73)
EVAL_RISK_ROWS  = [(118,125),(128,136),(139,147),(150,158)]
EVAL_COLS    = {"riesgo":2,"E":9,"P_":10,"C_":11,"barr_a":15,"barr_f":16,"jerarq":17,"epp":19,"E_r":20,"P_r":21,"C_r":22}
EVAL_DEFAULTS= {"E":1.2,"P_":1.2,"C_":3,"barr_a":"Charla de 5'. Definir medidas preventivas.","barr_f":"","jerarq":"CONTROLES ADMINISTRATIVOS","epp":"EPP Básicos","E_r":0.6,"P_r":0.6,"C_r":1.7}
COL = openpyxl.utils.column_index_from_string

def writable_anchors(ws, col, lo, hi):
    """Filas escribibles (anclas de merge) en un rango de una columna."""
    out = []
    for r in range(lo, hi + 1):
        if not isinstance(ws.cell(row=r, column=col), openpyxl.cell.cell.MergedCell):
            out.append(r)
    return out

def set_cell(ws, row, col, value):
    """Escribe respetando celdas combinadas (escribe en la celda ancla del rango)."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row) and (rng.min_col <= col <= rng.max_col):
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
    cell.value = value

def cargar_catalogo(path="catalogo_riesgos.json"):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def generar(spec, plantilla, salida, catalogo_path="catalogo_riesgos.json"):
    cat = cargar_catalogo(catalogo_path)
    shutil.copy(plantilla, salida)
    wb = openpyxl.load_workbook(salida, keep_vba=True)  # preserva macros y fórmulas

    # --- INSTALACIÓN ---
    inst = wb["INSTALACIÓN"]
    set_cell(inst, 11, COL("E"), spec["titulo"])
    set_cell(inst, 15, COL("E"), spec["instalacion"])
    set_cell(inst, 18, COL("E"), spec.get("version", 1))

    fases_ws = wb["FASES DE TAREA"]
    riesgos_ws = wb["RIESGOS"]

    for i, fase in enumerate(spec["fases"][:6]):
        # --- FASES DE TAREA ---
        trow = FASES_TITULO_ROW[i]
        set_cell(fases_ws, trow, COL("C"), fase["titulo"])
        set_cell(fases_ws, trow, COL("D"), fase.get("detalle", ""))
        set_cell(fases_ws, trow, COL("H"), fase.get("personas", ""))
        # herramientas apiladas en col I desde la fila del título
        for j, herr in enumerate(fase.get("herramientas", [])):
            set_cell(fases_ws, trow + j, COL("I"), herr)
        if fase.get("maquinaria"):
            set_cell(fases_ws, trow, COL("J"), fase["maquinaria"])
        if fase.get("izaje"):
            set_cell(fases_ws, trow, COL("L"), fase["izaje"])

        # --- RIESGOS (alimenta la matriz IPER vía VLOOKUP) ---
        lo, hi = RIESGOS_RANGE[i]
        anchors_d = writable_anchors(riesgos_ws, COL("D"), lo, hi)
        # limpiar riesgos y barreras viejos del rango de esta fase
        for r in anchors_d:
            set_cell(riesgos_ws, r, COL("D"), None)
            set_cell(riesgos_ws, r, COL("E"), None)
        # escribir los nuevos riesgos en los anclajes
        riesgos_fase = fase.get("riesgos", [])
        for k, riesgo in enumerate(riesgos_fase):
            if k >= len(anchors_d):
                break
            r = anchors_d[k]
            set_cell(riesgos_ws, r, COL("D"), riesgo)
            datos = cat.get(norm(riesgo))
            barrera = datos["barrera_admin"] if datos else "CONTROLES ADMINISTRATIVOS (charla de 5')"
            set_cell(riesgos_ws, r, COL("E"), barrera)
        # ocultar filas vacías (desde el primer slot sin datos hasta el final del rango)
        lo_r, hi_r = RIESGOS_RANGE[i]
        if len(riesgos_fase) < len(anchors_d):
            primer_vacio = anchors_d[len(riesgos_fase)]
            for r in range(primer_vacio, hi_r + 1):
                riesgos_ws.row_dimensions[r].hidden = True
        # asegurar que las filas usadas estén visibles
        for k in range(min(len(riesgos_fase), len(anchors_d))):
            r = anchors_d[k]
            riesgos_ws.row_dimensions[r].hidden = False

        # --- EVALUACION: sincronizar riesgos, Fine values y barreras ---
        if i < len(EVAL_RISK_ROWS):
            eval_ws = wb["EVALUACION"]
            ev_lo, ev_hi = EVAL_RISK_ROWS[i]
            for k in range(ev_hi - ev_lo + 1):
                r = ev_lo + k
                if k < len(riesgos_fase):
                    riesgo = riesgos_fase[k]
                    datos  = cat.get(norm(riesgo)) or {}
                    fi     = datos.get("fine_inherente") or {}
                    fr     = datos.get("fine_residual")  or {}
                    vals = {
                        "riesgo": riesgo,
                        "E":   fi.get("exp")  or EVAL_DEFAULTS["E"],
                        "P_":  fi.get("prob") or EVAL_DEFAULTS["P_"],
                        "C_":  fi.get("cons") or EVAL_DEFAULTS["C_"],
                        "barr_a": datos.get("barrera_admin") or EVAL_DEFAULTS["barr_a"],
                        "barr_f": datos.get("barrera_fisica") or EVAL_DEFAULTS["barr_f"],
                        "jerarq": datos.get("jerarquia")     or EVAL_DEFAULTS["jerarq"],
                        "epp":    datos.get("epp")           or EVAL_DEFAULTS["epp"],
                        "E_r":  fr.get("exp")  or EVAL_DEFAULTS["E_r"],
                        "P_r":  fr.get("prob") or EVAL_DEFAULTS["P_r"],
                        "C_r":  fr.get("cons") or EVAL_DEFAULTS["C_r"],
                    }
                    for key, col in EVAL_COLS.items():
                        set_cell(eval_ws, r, col, vals[key])
                    eval_ws.row_dimensions[r].hidden = False
                else:
                    # limpiar y ocultar filas sobrantes de esta fase en EVALUACION
                    for col in EVAL_COLS.values():
                        set_cell(eval_ws, r, col, None)
                    eval_ws.row_dimensions[r].hidden = True

    # Ocultar bloques de EVALUACION no usados
    num_fases = len(spec["fases"][:6])
    eval_ws = wb["EVALUACION"]
    if num_fases < len(EVAL_RISK_ROWS):
        ev_ocultar_desde = EVAL_RISK_ROWS[num_fases][0]
        for r in range(ev_ocultar_desde, eval_ws.max_row + 1):
            eval_ws.row_dimensions[r].hidden = True

    # Ocultar bloques de fases no usadas en FASES DE TAREA
    if num_fases < len(FASES_TITULO_ROW):
        primera_fila_libre = FASES_TITULO_ROW[num_fases]
        for r in range(primera_fila_libre, FASES_FIN_BLOQUE + 1):
            fases_ws.row_dimensions[r].hidden = True
    # Ocultar todo lo que sobra en RIESGOS desde el final del último bloque usado
    if num_fases < len(RIESGOS_RANGE):
        ocultar_desde = RIESGOS_RANGE[num_fases][0]
    else:
        ocultar_desde = RIESGOS_RANGE[-1][1] + 1
    for r in range(ocultar_desde, riesgos_ws.max_row + 1):
        riesgos_ws.row_dimensions[r].hidden = True

    wb.save(salida)
    return salida

if __name__ == "__main__":
    spec_path = sys.argv[1] if len(sys.argv) > 1 else "tarea_demo.json"
    plantilla = sys.argv[2] if len(sys.argv) > 2 else "plantilla.xlsm"
    salida = sys.argv[3] if len(sys.argv) > 3 else "AR_generado.xlsm"
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)
    out = generar(spec, plantilla, salida)
    print(f"AR generado -> {out}")
