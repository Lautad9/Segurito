"""Extrae el catálogo maestro (riesgo -> Fine + barreras) desde los AR .xlsm existentes."""
import openpyxl, warnings, json, sys, glob, unicodedata
from collections import defaultdict
warnings.filterwarnings("ignore")

def norm(s):
    """Normaliza nombre de riesgo: mayúsculas, sin tildes, sin dobles espacios, arregla typos."""
    if not s: return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = " ".join(s.split())
    fixes = {
        "PROYECCIONDE FRAGMENTOS O PARTICULAS": "PROYECCION DE FRAGMENTOS O PARTICULAS",
        "CAIDA DE PERSONAS AL DISTINTO NIVEL": "CAIDA DE PERSONAS A DISTINTO NIVEL",
        "ATRAPAMIENTO POR VUELCO DE MAQUINA": "ATRAPAMIENTO POR VUELCO DE MAQUINAS",
    }
    return fixes.get(s, s)

def extraer(files):
    cat = defaultdict(list)
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            print(f"  ! No pude abrir {f}: {e}"); continue
        if "EVALUACION" not in wb.sheetnames: continue
        ws = wb["EVALUACION"]
        for r in range(115, ws.max_row + 1):
            riesgo = norm(ws.cell(row=r, column=2).value)
            E = ws.cell(row=r, column=9).value   # Exposición inherente
            P = ws.cell(row=r, column=10).value  # Probabilidad
            C = ws.cell(row=r, column=11).value  # Consecuencia
            if not riesgo or riesgo in ("#REF!", "0") or riesgo.startswith("FASE"): continue
            if not all(isinstance(x, (int, float)) for x in (E, P, C)): continue
            cat[riesgo].append({
                "exp": E, "prob": P, "cons": C,
                "barrera_admin": str(ws.cell(row=r, column=15).value or "").strip(),
                "barrera_fisica": str(ws.cell(row=r, column=16).value or "").strip(),
                "jerarquia": str(ws.cell(row=r, column=17).value or "").strip(),
                "epp": str(ws.cell(row=r, column=19).value or "").strip(),
                "exp_res": ws.cell(row=r, column=20).value,
                "prob_res": ws.cell(row=r, column=21).value,
                "cons_res": ws.cell(row=r, column=22).value,
            })
    # Consolidar: por riesgo, tomar el valor de Fine más frecuente y la barrera más completa
    final = {}
    for riesgo, occ in cat.items():
        def moda(key):
            vals = [o[key] for o in occ if isinstance(o[key], (int, float))]
            return max(set(vals), key=vals.count) if vals else None
        mejor = max(occ, key=lambda o: len(o["barrera_admin"]) + len(o["barrera_fisica"]))
        final[riesgo] = {
            "fine_inherente": {"exp": moda("exp"), "prob": moda("prob"), "cons": moda("cons")},
            "fine_residual": {"exp": moda("exp_res"), "prob": moda("prob_res"), "cons": moda("cons_res")},
            "barrera_admin": mejor["barrera_admin"],
            "barrera_fisica": mejor["barrera_fisica"],
            "jerarquia": mejor["jerarquia"] or "CONTROLES ADMINISTRATIVOS - EPP",
            "epp": mejor["epp"] or "EPP Básicos",
            "ocurrencias": len(occ),
        }
    return final

if __name__ == "__main__":
    files = sys.argv[1:] or glob.glob("/mnt/user-data/uploads/*.xlsm")
    cat = extraer(files)
    out = "catalogo_riesgos.json"
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(cat, fh, ensure_ascii=False, indent=2)
    print(f"Catálogo con {len(cat)} riesgos -> {out}")
    for r, d in sorted(cat.items()):
        fi = d["fine_inherente"]
        print(f"  {r}: Fine E={fi['exp']} P={fi['prob']} C={fi['cons']} ({d['ocurrencias']} obs)")
