"""
Motor del backend. Tres pasos:
  1) rellenar_plantilla(spec) -> .xlsm con tu plantilla oficial completada
  2) recalcular(xlsm)        -> recalcula fórmulas (IPER, IMPRIMIR, scoring Fine) con LibreOffice
  3) exportar_pdf(xlsm)      -> PDF de la hoja IMPRIMIR, listo para firmar
"""
import os, shutil, subprocess, warnings, uuid
import openpyxl
from engine.generar_ar import generar  # reutiliza la lógica ya probada
warnings.filterwarnings("ignore")

BASE = os.path.dirname(os.path.abspath(__file__))
PLANTILLA = os.path.join(BASE, "plantilla.xlsm")
OUTDIR = os.environ.get("OUTPUTS_DIR", os.path.join(BASE, "outputs"))
if not os.path.isabs(OUTDIR) or OUTDIR == os.path.join(BASE, "outputs"):
    # En Railway usamos /tmp para persistencia dentro de la sesión
    if os.environ.get("RAILWAY_ENVIRONMENT"):
        OUTDIR = "/tmp/segurito_outputs"
CATALOGO = os.path.join(BASE, "engine", "catalogo_riesgos.json")
RECALC = os.path.join(BASE, "scripts", "recalc.py")
HOJA_IMPRIMIR = "IMPRIMIR "
os.makedirs(OUTDIR, exist_ok=True)

def _soffice(*args, timeout=180):
    """Llama a LibreOffice en modo headless."""
    cmd = ["soffice", "--headless", *args]
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)

def rellenar_plantilla(spec, job_id=None):
    job_id = job_id or uuid.uuid4().hex[:10]
    out_xlsm = os.path.join(OUTDIR, f"AR_{job_id}.xlsm")
    generar(spec, PLANTILLA, out_xlsm, catalogo_path=CATALOGO)
    return out_xlsm, job_id

def recalcular(xlsm, timeout=180):
    """Recalcula todas las fórmulas y devuelve el conteo de errores."""
    if os.path.exists(RECALC):
        r = subprocess.run(["python3", RECALC, xlsm, str(timeout)],
                           capture_output=True, text=True, timeout=timeout + 30)
        return r.stdout
    return None

def filtrar_pdf(pdf_path):
    """Elimina páginas con #¡VALOR! / #VALUE! del PDF generado."""
    import pypdf
    reader = pypdf.PdfReader(pdf_path)
    writer = pypdf.PdfWriter()
    removidas = 0
    for i, page in enumerate(reader.pages):
        txt = page.extract_text() or ""
        if "#¡VALOR!" in txt or "#VALUE!" in txt:
            removidas += 1
        else:
            writer.add_page(page)
    if removidas > 0:
        with open(pdf_path, "wb") as f:
            writer.write(f)
        print(f"[PDF] {removidas} páginas con errores eliminadas → {len(reader.pages)-removidas} páginas finales")

def exportar_pdf(xlsm, job_id):
    """Deja visible solo la hoja IMPRIMIR y exporta a PDF."""
    tmp = os.path.join(OUTDIR, f"_pdf_{job_id}.xlsm")
    shutil.copy(xlsm, tmp)
    wb = openpyxl.load_workbook(tmp, keep_vba=True)
    for ws in wb.worksheets:
        ws.sheet_state = "visible" if ws.title == HOJA_IMPRIMIR else "hidden"
    if HOJA_IMPRIMIR in wb.sheetnames:
        wb.active = wb.sheetnames.index(HOJA_IMPRIMIR)
    wb.save(tmp)
    _soffice("--convert-to", "pdf", "--outdir", OUTDIR, tmp)
    pdf_tmp = tmp.replace(".xlsm", ".pdf")
    pdf_out = os.path.join(OUTDIR, f"AR_{job_id}.pdf")
    if os.path.exists(pdf_tmp):
        os.replace(pdf_tmp, pdf_out)
    if os.path.exists(tmp):
        os.remove(tmp)
    if os.path.exists(pdf_out):
        filtrar_pdf(pdf_out)
    return pdf_out if os.path.exists(pdf_out) else None

def generar_todo(spec):
    """Pipeline completo: rellena + recalcula + PDF. Devuelve rutas."""
    xlsm, job_id = rellenar_plantilla(spec)
    recalcular(xlsm)
    pdf = exportar_pdf(xlsm, job_id)
    return {"job_id": job_id, "xlsm": xlsm, "pdf": pdf}
